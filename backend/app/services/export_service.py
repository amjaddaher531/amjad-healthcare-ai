"""
Report export: JSON is just the Pydantic model dump; PDF uses reportlab;
Excel uses openpyxl with one sheet per major table.
"""
import io
from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl

from app.models.schemas import PipelineResult


def build_pdf(result: PipelineResult) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LETTER, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], textColor=colors.HexColor("#0B3D3A"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#0B3D3A"))

    fr = result.final_report
    story = [
        Paragraph("Amjad Healthcare AI — Medical Coding &amp; Audit Report", title_style),
        Paragraph(f"Case ID: {fr.case_id}  |  Status: {fr.final_status.value}  |  Confidence: {fr.confidence_score:.1f}%", styles["Normal"]),
        Spacer(1, 12),
        Paragraph("Executive Summary", h2),
        Paragraph(fr.executive_summary, styles["Normal"]),
        Spacer(1, 12),
    ]

    def code_table(title: str, rows):
        story.append(Paragraph(title, h2))
        if not rows:
            story.append(Paragraph("None identified.", styles["Normal"]))
            return
        data = [["Code", "Description", "Confidence", "Risk"]]
        for c in rows:
            data.append([c.code, c.description[:60], f"{c.confidence_score:.0f}%", c.risk_level.value])
        t = Table(data, colWidths=[70, 260, 70, 60])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D3A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    code_table("Diagnosis Codes (ICD-10-CM)", fr.diagnosis_table)
    code_table("Procedure Codes (CPT)", fr.procedure_table)
    code_table("HCPCS Codes", fr.hcpcs_table)
    code_table("Modifiers", fr.modifier_table)

    story.append(Paragraph("Compliance Review", h2))
    if fr.compliance_review.issues:
        for issue in fr.compliance_review.issues:
            story.append(Paragraph(f"[{issue.severity.value}] {issue.category}: {issue.description}", styles["Normal"]))
    else:
        story.append(Paragraph("No compliance issues identified.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Recommendations", h2))
    for rec in fr.recommendations:
        story.append(Paragraph(f"• {rec}", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def build_excel(result: PipelineResult) -> bytes:
    wb = openpyxl.Workbook()
    fr = result.final_report

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Case ID", fr.case_id])
    ws.append(["Final Status", fr.final_status.value])
    ws.append(["Confidence Score", fr.confidence_score])
    ws.append(["Executive Summary", fr.executive_summary])

    def sheet_for(name: str, rows):
        s = wb.create_sheet(name)
        s.append(["Code", "Type", "Description", "Reason", "Evidence", "Confidence", "Risk"])
        for c in rows:
            s.append([c.code, c.code_type, c.description, c.reason, c.supporting_evidence,
                      c.confidence_score, c.risk_level.value])

    sheet_for("ICD-10-CM", fr.diagnosis_table)
    sheet_for("CPT", fr.procedure_table)
    sheet_for("HCPCS", fr.hcpcs_table)
    sheet_for("Modifiers", fr.modifier_table)

    compliance_sheet = wb.create_sheet("Compliance")
    compliance_sheet.append(["Category", "Severity", "Description", "Related Code", "Recommendation"])
    for issue in fr.compliance_review.issues:
        compliance_sheet.append([issue.category, issue.severity.value, issue.description,
                                  issue.related_code or "", issue.recommendation])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
